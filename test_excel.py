import openpyxl
from numpy.ma.extras import row_stack


def read_excel_file(file_name, sheet_name: str = None):
    wb = openpyxl.load_workbook(file_name)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    rows = ws.max_row
    columns = ws.max_column
    print(f"The file has {rows} rows and {columns} columns.")
    return [i for i in ws.iter_rows(values_only=True)]


excel_file_operations_prompt = """
You are an agent responsible for performing operations on Excel files.

# Primary Instructions

"""

def add_data_to_excel_file(file_name, sheet_name, data):
    # insert a new row at the end of the sheet
    wb = openpyxl.load_workbook(file_name)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    for row in data:
        ws.append(row)
    wb.save(file_name)


# data = read_excel_file("/home/kevit/Downloads/Aircotedivoire Test.xlsx")
# column_names = data[0]
# row_data = data[1:]

# add_data_to_excel_file("/home/kevit/Downloads/Aircotedivoire Test.xlsx", "Sheet1", [["kevit", "hetvee", "shah"]])